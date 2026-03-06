import argparse
import sys
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

REQUIRED_COLUMNS = {"client_id", "date", "amount"}
DEFAULT_THRESHOLD = 2.5
DEFAULT_MIN_RECORDS = 5


def load_transactions(filepath: str) -> pd.DataFrame:
    path = Path(filepath)
    if not path.exists():
        print(f"Error: file not found: {filepath}")
        sys.exit(1)

    if path.suffix == ".csv":
        df = pd.read_csv(filepath, parse_dates=["date"])
    elif path.suffix in (".xlsx", ".xls"):
        df = pd.read_excel(filepath, parse_dates=["date"])
    else:
        print("Error: input must be .csv or .xlsx")
        sys.exit(1)

    missing = REQUIRED_COLUMNS - set(df.columns.str.lower())
    if missing:
        print(f"Error: missing required columns: {missing}")
        sys.exit(1)

    df.columns = df.columns.str.lower()
    return df


def compute_z_scores(df: pd.DataFrame, min_records: int) -> pd.DataFrame:
    """
    Compute z-scores per client for transaction amounts.
    Clients with less than min_records of transactions will be ignored
    to prevent poor standard deviation estimates.
    """
    df = df.copy()
    df["z_score"] = np.nan

    grouped = df.groupby("client_id")["amount"]
    means = grouped.transform("mean")
    stds = grouped.transform("std")
    counts = grouped.transform("count")

    # only score clients with enough history
    sufficient = counts >= min_records
    df.loc[sufficient, "z_score"] = (
        (df.loc[sufficient, "amount"] - means[sufficient]) / stds[sufficient]
    )
    return df


def flag_outliers(df: pd.DataFrame, threshold: float) -> pd.DataFrame:
    flagged = df[df["z_score"].abs() > threshold].copy()
    flagged["severity"] = flagged["z_score"].abs().apply(
        lambda z: "HIGH" if z > threshold * 1.5 else "MEDIUM"
    )
    return flagged.sort_values("z_score", key=abs, ascending=False)


def export_report(flagged: pd.DataFrame, output_path: str) -> None:
    flagged.to_excel(output_path, index=False)

    wb = load_workbook(output_path)
    ws = wb.active

    high_fill = PatternFill(start_color="FF4C4C", end_color="FF4C4C", fill_type="solid")
    med_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    header_font = Font(bold=True)

    # bold header row
    for cell in ws[1]:
        cell.font = header_font

    # color rows by severity
    severity_col = [cell.value for cell in ws[1]].index("severity") + 1
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        severity = row[severity_col - 1].value
        fill = high_fill if severity == "HIGH" else med_fill
        for cell in row:
            cell.fill = fill

    wb.save(output_path)


def print_summary(df: pd.DataFrame, flagged: pd.DataFrame, threshold: float) -> None:
    total = len(df)
    n_flagged = len(flagged)
    n_skipped = df["z_score"].isna().sum()

    print(f"\n{'='*50}")
    print(f"  Transaction Anomaly Detection Report")
    print(f"{'='*50}")
    print(f"  Total records:        {total}")
    print(f"  Skipped (low volume): {n_skipped}")
    print(f"  Threshold (|z|):      {threshold}")
    print(f"  Flagged:              {n_flagged}")
    if n_flagged:
        high = (flagged["severity"] == "HIGH").sum()
        med = (flagged["severity"] == "MEDIUM").sum()
        print(f"    HIGH severity:      {high}")
        print(f"    MEDIUM severity:    {med}")
    print(f"{'='*50}\n")


def main():
    parser = argparse.ArgumentParser(
        description="Flag statistically unusual clients based upon their financial transaction data using z-scores."
    )
    parser.add_argument("input", help="Path to input file (.csv or .xlsx)")
    parser.add_argument(
        "-o", "--output", default="flagged_transactions.xlsx",
        help="Output Excel report path (default: flagged_transactions.xlsx)"
    )
    parser.add_argument(
        "-t", "--threshold", type=float, default=DEFAULT_THRESHOLD,
        help=f"Z-score threshold to use for flagging (default = {DEFAULT_THRESHOLD})."
    )
    parser.add_argument(
        "--min-records", type=int, default=DEFAULT_MIN_RECORDS,
        help=f"Minimum number of transactions required per client for inclusion in scoring (default = {DEFAULT_MIN_RECORDS})."
    )
    args = parser.parse_args()

    df = load_transactions(args.input)
    df = compute_z_scores(df, args.min_records)
    flagged = flag_outliers(df, args.threshold)

    print_summary(df, flagged, args.threshold)

    if not flagged.empty:
        export_report(flagged, args.output)
        print(f"Report saved to: {args.output}")
    else:
        print("No anomalies detected above threshold.")


if __name__ == "__main__":
    main()
